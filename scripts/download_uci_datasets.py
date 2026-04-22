#!/usr/bin/env python3
"""Download and prepare UCI datasets for E2 benchmark.

Downloads Power, Gas, HEPMASS, and MiniBoone datasets and saves them
in the expected format for vdc.data.tabular.maybe_load_uci().
"""

import argparse
import sys
import urllib.request
import gzip
from pathlib import Path
import tarfile
import zipfile

import numpy as np

REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from vdc.data.paths import data_root

# Dataset URLs and configurations
UCI_DATASETS = {
    'power': {
        'url': 'https://archive.ics.uci.edu/ml/machine-learning-databases/00294/CCPP.zip',
        'extract': 'zip',
        'file_in_archive': 'CCPP/Folds5x2_pp.xlsx',
        'preprocess': 'power',
    },
    'gas': {
        'url': 'https://archive.ics.uci.edu/ml/machine-learning-databases/00322/data.zip',
        'extract': 'zip',
        'preprocess': 'gas',
    },
    'hepmass': {
        'url': 'https://archive.ics.uci.edu/ml/machine-learning-databases/00347/all_train.csv.gz',
        'extract': 'gzip',
        'preprocess': 'hepmass',
    },
    'miniboone': {
        'url': 'https://archive.ics.uci.edu/ml/machine-learning-databases/00199/MiniBooNE_PID.txt',
        'extract': None,
        'preprocess': 'miniboone',
    },
}


def download_file(url: str, dest: Path) -> Path:
    """Download a file from URL to destination."""
    print(f"  Downloading {url}")
    print(f"  -> {dest}")
    dest.parent.mkdir(parents=True, exist_ok=True)
    urllib.request.urlretrieve(url, dest)
    return dest


def preprocess_power(raw_dir: Path, out_dir: Path):
    """Preprocess Power dataset."""
    xlsx_path = raw_dir / 'CCPP' / 'Folds5x2_pp.xlsx'
    if not xlsx_path.exists():
        matches = list(raw_dir.glob('**/Folds5x2_pp.xlsx'))
        xlsx_path = matches[0] if matches else None
    
    data = None
    
    # Try openpyxl first (more reliable)
    if xlsx_path and xlsx_path.exists():
        try:
            from openpyxl import load_workbook
            wb = load_workbook(xlsx_path)
            ws = wb.active
            data = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] is not None:
                    data.append([float(x) for x in row])
            data = np.array(data, dtype=np.float32)
            print(f"  Loaded Power data via openpyxl: {data.shape}")
        except ImportError:
            pass
        except Exception as e:
            print(f"  openpyxl failed: {e}")
    
    # Try pandas as fallback
    if data is None and xlsx_path and xlsx_path.exists():
        try:
            import pandas as pd
            df = pd.read_excel(xlsx_path, engine='openpyxl')
            data = df.values.astype(np.float32)
            print(f"  Loaded Power data via pandas: {data.shape}")
        except Exception as e:
            print(f"  pandas failed: {e}")
    
    if data is None:
        raise RuntimeError(
            "Could not load the Power dataset from the downloaded archive. "
            "Install openpyxl and verify the archive contents."
        )
    
    # Split into train/test (80/20)
    n = len(data)
    np.random.seed(42)
    idx = np.random.permutation(n)
    split = int(0.8 * n)
    train = data[idx[:split]]
    test = data[idx[split:]]
    
    out_dir.mkdir(parents=True, exist_ok=True)
    np.savez(out_dir / 'power.npz', train=train, test=test)
    print(f"  Saved power.npz: train {train.shape}, test {test.shape}")
    return True


def preprocess_gas(raw_dir: Path, out_dir: Path):
    """Preprocess Gas sensor dataset."""
    # Find the data file
    data_files = list(raw_dir.glob('**/*.dat')) + list(raw_dir.glob('**/*.txt'))
    
    all_data = []
    for f in sorted(raw_dir.glob('**/batch*.dat')):
        with open(f, 'r') as fh:
            for line in fh:
                parts = line.strip().split()
                if len(parts) > 1:
                    # Format: class feature:value feature:value ...
                    features = []
                    for p in parts[1:]:
                        if ':' in p:
                            features.append(float(p.split(':')[1]))
                    if features:
                        all_data.append(features)
    
    if not all_data:
        # Try alternative format
        for f in raw_dir.glob('**/*.csv'):
            try:
                data = np.loadtxt(f, delimiter=',', skiprows=1)
                all_data.extend(data.tolist())
            except:
                pass
    
    if not all_data:
        raise RuntimeError("Could not parse the Gas dataset after download and extraction.")

    data = np.array(all_data, dtype=np.float32)
    n = len(data)
    np.random.seed(42)
    idx = np.random.permutation(n)
    split = int(0.8 * n)
    train = data[idx[:split]]
    test = data[idx[split:]]
    
    out_dir.mkdir(parents=True, exist_ok=True)
    np.savez(out_dir / 'gas.npz', train=train, test=test)
    print(f"  Saved gas.npz: train {train.shape}, test {test.shape}")
    return True


def preprocess_hepmass(raw_dir: Path, out_dir: Path):
    """Preprocess HEPMASS dataset."""
    csv_path = raw_dir / 'all_train.csv'
    if not csv_path.exists():
        csv_path = list(raw_dir.glob('*.csv'))[0] if list(raw_dir.glob('*.csv')) else None
    
    if csv_path and csv_path.exists():
        # Load CSV, skip header
        data = np.loadtxt(csv_path, delimiter=',', skiprows=1, max_rows=500000)
        # Remove the label column (first column) and mass column
        data = data[:, 1:]  # Remove label
        if data.shape[1] > 21:
            # Remove mass column (column 0 after removing label)
            data = np.delete(data, 0, axis=1)
    else:
        raise RuntimeError("Could not find the HEPMASS CSV after download and extraction.")
    
    data = data.astype(np.float32)
    n = len(data)
    np.random.seed(42)
    idx = np.random.permutation(n)
    split = int(0.8 * n)
    train = data[idx[:split]]
    test = data[idx[split:]]
    
    out_dir.mkdir(parents=True, exist_ok=True)
    np.savez(out_dir / 'hepmass.npz', train=train, test=test)
    print(f"  Saved hepmass.npz: train {train.shape}, test {test.shape}")
    return True


def preprocess_miniboone(raw_dir: Path, out_dir: Path):
    """Preprocess MiniBooNE dataset."""
    txt_path = raw_dir / 'MiniBooNE_PID.txt'
    if not txt_path.exists():
        txt_path = list(raw_dir.glob('*.txt'))[0] if list(raw_dir.glob('*.txt')) else None
    
    if txt_path and txt_path.exists():
        with open(txt_path, 'r') as f:
            first_line = f.readline().strip().split()
            n_signal = int(first_line[0])
            n_background = int(first_line[1])
        
        data = np.loadtxt(txt_path, skiprows=1)
        # Use signal events only for cleaner distribution
        data = data[:n_signal]
    else:
        raise RuntimeError("Could not find the MiniBooNE data after download.")
    
    data = data.astype(np.float32)
    n = len(data)
    np.random.seed(42)
    idx = np.random.permutation(n)
    split = int(0.8 * n)
    train = data[idx[:split]]
    test = data[idx[split:]]
    
    out_dir.mkdir(parents=True, exist_ok=True)
    np.savez(out_dir / 'miniboone.npz', train=train, test=test)
    print(f"  Saved miniboone.npz: train {train.shape}, test {test.shape}")
    return True


def download_and_prepare(dataset: str, output_base: Path):
    """Download and prepare a single dataset."""
    print(f"\n{'='*60}")
    print(f"Processing: {dataset}")
    print('='*60)
    
    cfg = UCI_DATASETS[dataset]
    raw_dir = output_base / 'raw' / dataset
    out_dir = output_base / 'uci'
    raw_dir.mkdir(parents=True, exist_ok=True)
    
    # Download
    url = cfg['url']
    ext = url.split('.')[-1]
    if cfg['extract'] == 'gzip':
        dl_path = raw_dir / f'{dataset}.csv.gz'
    elif cfg['extract'] == 'zip':
        dl_path = raw_dir / f'{dataset}.zip'
    else:
        dl_path = raw_dir / url.split('/')[-1]
    
    if not dl_path.exists():
        try:
            download_file(url, dl_path)
        except Exception as e:
            raise RuntimeError(f"Download failed for {dataset}: {e}") from e
    
    # Extract
    if cfg['extract'] == 'zip':
        print(f"  Extracting ZIP...")
        with zipfile.ZipFile(dl_path, 'r') as zf:
            zf.extractall(raw_dir)
    elif cfg['extract'] == 'gzip':
        print(f"  Extracting GZIP...")
        csv_path = raw_dir / f'{dataset}.csv'
        with gzip.open(dl_path, 'rb') as f_in:
            with open(csv_path, 'wb') as f_out:
                f_out.write(f_in.read())
    
    # Preprocess
    print(f"  Preprocessing...")
    preprocess_fn = {
        'power': preprocess_power,
        'gas': preprocess_gas,
        'hepmass': preprocess_hepmass,
        'miniboone': preprocess_miniboone,
    }[cfg['preprocess']]
    
    return preprocess_fn(raw_dir, out_dir)


def main():
    parser = argparse.ArgumentParser(description='Download UCI datasets')
    parser.add_argument(
        '--output-base',
        type=str,
        default=None,
        help='Directory that will contain raw/ and uci/ subdirectories (defaults to DATA_ROOT or repo-local data/).',
    )
    parser.add_argument('--datasets', nargs='+', 
                        default=['power', 'gas', 'hepmass', 'miniboone'],
                        help='Datasets to download')
    args = parser.parse_args()
    
    output_base = Path(args.output_base).expanduser() if args.output_base else data_root()
    print(f"Output base: {output_base}")

    failures = []
    for ds in args.datasets:
        if ds not in UCI_DATASETS:
            print(f"Unknown dataset: {ds}")
            continue
        try:
            download_and_prepare(ds, output_base)
        except Exception as exc:
            failures.append((ds, exc))
            print(f"Failed to prepare {ds}: {exc}")
    
    print("\n" + "="*60)
    if failures:
        print("Completed with failures.")
        for ds, exc in failures:
            print(f"  {ds}: {exc}")
        print("="*60)
        raise SystemExit(1)
    print("Done! Datasets saved to:", output_base / 'uci')
    print("="*60)


if __name__ == '__main__':
    main()
